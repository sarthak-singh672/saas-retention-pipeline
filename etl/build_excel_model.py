# =============================================================
# B2B SaaS Retention Pipeline — Excel Model Builder
# File: etl/build_excel_model.py
# Purpose: Build professional Excel workbook with 4 sheets
# =============================================================

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
import os

print("=" * 60)
print("B2B SaaS Retention Pipeline — Excel Model Builder")
print("=" * 60)

# -------------------------------------------------------------
# SECTION 0: Load All Data
# -------------------------------------------------------------
print("\n[0/5] Loading processed data...")

accounts    = pd.read_csv('processed/clean_accounts.csv')
subs        = pd.read_csv('processed/clean_subscriptions.csv')
churn       = pd.read_csv('processed/clean_churn_events.csv')
tickets     = pd.read_csv('processed/clean_support_tickets.csv')
at_risk     = pd.read_csv('processed/result_retention_recommendations.csv')
mrr_trends  = pd.read_csv('processed/result_mrr_trends.csv')
ltv         = pd.read_csv('processed/result_ltv.csv')

print("  All files loaded successfully ✓")

# -------------------------------------------------------------
# Helper: Styling
# -------------------------------------------------------------
# Colors
NAVY        = "1B3A6B"
TEAL        = "0E7C7B"
LIGHT_BLUE  = "D6E4F0"
LIGHT_GREEN = "D5F5E3"
LIGHT_RED   = "FADBD8"
YELLOW      = "FEF9E7"
WHITE       = "FFFFFF"
DARK_GRAY   = "2C3E50"
MID_GRAY    = "BDC3C7"

def header_style(cell, bg=NAVY, font_color=WHITE, bold=True, size=11):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=bold, color=font_color, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)

def data_style(cell, bg=WHITE, bold=False, align="center"):
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=bold, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")

def thin_border(cell):
    thin = Side(style="thin", color="BDC3C7")
    cell.border = Border(left=thin, right=thin,
                         top=thin, bottom=thin)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def merge_title(ws, cell_range, text, bg=NAVY, size=14):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value     = text
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.font      = Font(bold=True, color=WHITE, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Create Workbook
wb = Workbook()
wb.remove(wb.active)   # remove default blank sheet


# =============================================================
# SHEET 1: SLA MODEL
# =============================================================
print("\n[1/5] Building SLA Model sheet...")

ws1 = wb.create_sheet("SLA Model")
ws1.sheet_view.showGridLines = False
ws1.row_dimensions[1].height = 35
ws1.row_dimensions[2].height = 20

# Title
merge_title(ws1, "A1:H1",
    "SLA Performance Model — Support Response & Resolution Analysis")

# --- Section A: SLA Targets ---
ws1["A3"] = "SLA TARGETS BY PRIORITY"
ws1["A3"].font = Font(bold=True, color=NAVY, size=12)

headers = ["Priority", "Target Response (mins)",
           "Target Resolution (hrs)", "Penalty if Breached"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=4, column=col, value=h)
    header_style(c, bg=TEAL)
    set_col_width(ws1, col, 22)

sla_targets = [
    ("HIGH",    15,  4,  "Credit 10% of monthly fee"),
    ("MEDIUM",  60,  24, "Credit 5% of monthly fee"),
    ("LOW",     240, 72, "No financial penalty"),
]
for r, (pri, resp, res, pen) in enumerate(sla_targets, 5):
    bg = LIGHT_RED if pri == "HIGH" else (
         YELLOW if pri == "MEDIUM" else LIGHT_GREEN)
    vals = [pri, resp, res, pen]
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=c, value=v)
        data_style(cell, bg=bg)
        thin_border(cell)

# --- Section B: Actual SLA Performance ---
ws1["A9"] = "ACTUAL SLA PERFORMANCE"
ws1["A9"].font = Font(bold=True, color=NAVY, size=12)

# Calculate from tickets data
sla_actual = (
    tickets.groupby('priority').agg(
        total_tickets      = ('ticket_id',                  'count'),
        avg_response_mins  = ('first_response_time_minutes','mean'),
        avg_resolution_hrs = ('resolution_time_hours',      'mean'),
        sla_breaches       = ('sla_breached',               'sum'),
        avg_satisfaction   = ('satisfaction_score',         'mean')
    ).reset_index()
)

perf_headers = [
    "Priority", "Total Tickets", "Avg Response (mins)",
    "Avg Resolution (hrs)", "SLA Breaches", "Avg Satisfaction"
]
for col, h in enumerate(perf_headers, 1):
    c = ws1.cell(row=10, column=col, value=h)
    header_style(c, bg=DARK_GRAY)

for r, row in enumerate(sla_actual.itertuples(), 11):
    vals = [
        row.priority,
        row.total_tickets,
        round(row.avg_response_mins, 1),
        round(row.avg_resolution_hrs, 1),
        row.sla_breaches,
        round(row.avg_satisfaction, 2)
    ]
    bg = LIGHT_RED if row.sla_breaches > 10 else LIGHT_GREEN
    for c, v in enumerate(vals, 1):
        cell = ws1.cell(row=r, column=c, value=v)
        data_style(cell, bg=bg)
        thin_border(cell)

# --- Section C: SLA Summary KPIs ---
ws1["A16"] = "SLA SUMMARY KPIs"
ws1["A16"].font = Font(bold=True, color=NAVY, size=12)

total_breaches   = int(tickets['sla_breached'].sum())
breach_rate      = round(tickets['sla_breached'].mean() * 100, 1)
avg_satisfaction = round(tickets['satisfaction_score'].mean(), 2)
escalation_rate  = round(tickets['escalation_flag'].mean() * 100, 1)

kpis = [
    ("Total SLA Breaches",      total_breaches,   "count"),
    ("SLA Breach Rate",         f"{breach_rate}%","percent"),
    ("Avg Satisfaction Score",  avg_satisfaction, "score"),
    ("Escalation Rate",         f"{escalation_rate}%","percent"),
]
for col, (label, value, _) in enumerate(kpis, 1):
    ws1.cell(row=17, column=col, value=label).font = Font(
        bold=True, color=DARK_GRAY)
    cell = ws1.cell(row=18, column=col, value=value)
    data_style(cell, bg=LIGHT_BLUE, bold=True)
    thin_border(cell)

print("  SLA Model sheet complete ✓")


# =============================================================
# SHEET 2: REVENUE SCENARIO GRID
# =============================================================
print("\n[2/5] Building Revenue Scenario Grid sheet...")

ws2 = wb.create_sheet("Revenue Scenario Grid")
ws2.sheet_view.showGridLines = False
ws2.row_dimensions[1].height = 35

merge_title(ws2, "A1:I1",
    "Revenue Scenario Grid — Impact of Churn Reduction on MRR")

# Base metrics
total_accounts  = len(accounts)
churned_accounts= int(accounts['churn_flag'].sum())
current_mrr     = round(subs['mrr_amount'].sum(), 2)
avg_mrr_account = round(current_mrr / total_accounts, 2)
current_churn   = round(churned_accounts / total_accounts * 100, 1)

# Base metrics display
ws2["A3"] = "BASE METRICS"
ws2["A3"].font = Font(bold=True, color=NAVY, size=12)

base_data = [
    ("Total Accounts",          total_accounts),
    ("Churned Accounts",        churned_accounts),
    ("Current Churn Rate",      f"{current_churn}%"),
    ("Total MRR",               f"${current_mrr:,.2f}"),
    ("Avg MRR per Account",     f"${avg_mrr_account:,.2f}"),
]
for col, (label, value) in enumerate(base_data, 1):
    ws2.cell(row=4, column=col,
             value=label).font = Font(bold=True, color=DARK_GRAY)
    cell = ws2.cell(row=5, column=col, value=value)
    data_style(cell, bg=LIGHT_BLUE, bold=True)
    thin_border(cell)
    set_col_width(ws2, col, 20)

# Scenario Grid
ws2["A8"] = "CHURN REDUCTION SCENARIO GRID"
ws2["A8"].font = Font(bold=True, color=NAVY, size=12)

scenario_headers = [
    "Scenario", "Churn Reduction", "New Churn Rate",
    "Accounts Saved", "MRR Recovered", "ARR Impact",
    "3-Year Revenue Gain", "ROI vs Retention Cost", "Recommendation"
]
for col, h in enumerate(scenario_headers, 1):
    c = ws2.cell(row=9, column=col, value=h)
    header_style(c, bg=NAVY)
    set_col_width(ws2, col, 20)

scenarios = [
    ("Pessimistic",  5,  "Monitor only"),
    ("Conservative", 10, "Basic CSM outreach"),
    ("Moderate",     15, "Targeted campaigns"),
    ("Optimistic",   20, "Full retention programme"),
    ("Aggressive",   30, "Executive-led retention"),
]

retention_cost = 50000   # assumed annual retention programme cost

for r, (name, reduction_pct, action) in enumerate(scenarios, 10):
    accounts_saved   = round(churned_accounts * reduction_pct / 100)
    mrr_recovered    = round(accounts_saved * avg_mrr_account, 2)
    arr_impact       = round(mrr_recovered * 12, 2)
    three_yr_gain    = round(arr_impact * 3, 2)
    new_churn_rate   = round(current_churn * (1 - reduction_pct/100), 1)
    roi              = round((three_yr_gain - retention_cost) /
                              retention_cost * 100, 1)

    bg = (LIGHT_GREEN if reduction_pct >= 20 else
          YELLOW      if reduction_pct >= 10 else
          LIGHT_RED)

    vals = [
        name,
        f"{reduction_pct}%",
        f"{new_churn_rate}%",
        accounts_saved,
        f"${mrr_recovered:,.2f}",
        f"${arr_impact:,.2f}",
        f"${three_yr_gain:,.2f}",
        f"{roi}%",
        action
    ]
    for col, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=col, value=v)
        data_style(cell, bg=bg)
        thin_border(cell)

print("  Revenue Scenario Grid sheet complete ✓")


# =============================================================
# SHEET 3: CHURN SUMMARY
# =============================================================
print("\n[3/5] Building Churn Summary sheet...")

ws3 = wb.create_sheet("Churn Summary")
ws3.sheet_view.showGridLines = False
ws3.row_dimensions[1].height = 35

merge_title(ws3, "A1:G1",
    "Churn Summary — Financial Impact by Plan Tier")

# Churn by plan
churn_by_plan = (
    accounts.groupby('plan_tier').agg(
        total    = ('account_id', 'count'),
        churned  = ('churn_flag', 'sum')
    ).reset_index()
)
churn_by_plan['retained']   = (churn_by_plan['total'] -
                                churn_by_plan['churned'])
churn_by_plan['churn_rate'] = (churn_by_plan['churned'] /
                                churn_by_plan['total'] * 100).round(1)

# MRR by plan
mrr_by_plan = (
    subs.groupby('plan_tier')['mrr_amount']
    .sum().reset_index()
    .rename(columns={'mrr_amount': 'total_mrr'})
)
churn_by_plan = churn_by_plan.merge(mrr_by_plan,
                                     on='plan_tier', how='left')
churn_by_plan['mrr_at_risk'] = (
    churn_by_plan['total_mrr'] *
    churn_by_plan['churn_rate'] / 100
).round(2)

headers = [
    "Plan Tier", "Total Accounts", "Churned",
    "Retained", "Churn Rate %", "Total MRR", "MRR at Risk"
]
for col, h in enumerate(headers, 1):
    c = ws3.cell(row=3, column=col, value=h)
    header_style(c, bg=NAVY)
    set_col_width(ws3, col, 18)

for r, row in enumerate(churn_by_plan.itertuples(), 4):
    bg = (LIGHT_RED   if row.churn_rate > 30 else
          YELLOW      if row.churn_rate > 20 else
          LIGHT_GREEN)
    vals = [
        row.plan_tier,
        row.total,
        row.churned,
        row.retained,
        f"{row.churn_rate}%",
        f"${row.total_mrr:,.2f}",
        f"${row.mrr_at_risk:,.2f}"
    ]
    for col, v in enumerate(vals, 1):
        cell = ws3.cell(row=r, column=col, value=v)
        data_style(cell, bg=bg)
        thin_border(cell)

# Totals row
total_row = len(churn_by_plan) + 4
total_vals = [
    "TOTAL",
    churn_by_plan['total'].sum(),
    churn_by_plan['churned'].sum(),
    churn_by_plan['retained'].sum(),
    f"{round(churn_by_plan['churned'].sum() / churn_by_plan['total'].sum() * 100, 1)}%",
    f"${churn_by_plan['total_mrr'].sum():,.2f}",
    f"${churn_by_plan['mrr_at_risk'].sum():,.2f}"
]
for col, v in enumerate(total_vals, 1):
    cell = ws3.cell(row=total_row, column=col, value=v)
    header_style(cell, bg=DARK_GRAY)
    thin_border(cell)

print("  Churn Summary sheet complete ✓")


# =============================================================
# SHEET 4: RETENTION DASHBOARD
# =============================================================
print("\n[4/5] Building Retention Dashboard sheet...")

ws4 = wb.create_sheet("Retention Dashboard")
ws4.sheet_view.showGridLines = False
ws4.row_dimensions[1].height = 35

merge_title(ws4, "A1:J1",
    "Retention Action Dashboard — At-Risk Customer Recommendations")

headers = [
    "Account", "Plan", "Industry", "Country",
    "Churn %", "Priority", "Primary Action",
    "Offer", "Key Risk", "Success Metric"
]
for col, h in enumerate(headers, 1):
    c = ws4.cell(row=3, column=col, value=h)
    header_style(c, bg=TEAL)

col_widths = [18, 12, 15, 12, 10, 16, 35, 30, 30, 30]
for col, width in enumerate(col_widths, 1):
    set_col_width(ws4, col, width)

for r, row in enumerate(at_risk.itertuples(), 4):
    bg = (LIGHT_RED   if row.churn_probability >= 40 else
          YELLOW      if row.churn_probability >= 25 else
          LIGHT_GREEN)

    vals = [
        row.account_name,
        row.plan_tier,
        row.industry,
        row.country,
        f"{row.churn_probability}%",
        row.priority_action,
        row.primary_action,
        row.offer,
        row.key_risk_factor,
        row.success_metric
    ]
    for col, v in enumerate(vals, 1):
        cell = ws4.cell(row=r, column=col, value=v)
        align = "left" if col >= 7 else "center"
        data_style(cell, bg=bg, align=align)
        thin_border(cell)
        ws4.row_dimensions[r].height = 30

print("  Retention Dashboard sheet complete ✓")


# =============================================================
# SAVE WORKBOOK
# =============================================================
print("\n[5/5] Saving Excel workbook...")

output_path = 'excel/saas_retention_model.xlsx'
os.makedirs('excel', exist_ok=True)
wb.save(output_path)

print("\n" + "=" * 60)
print("EXCEL MODEL — Complete!")
print("=" * 60)
print(f"  Saved to : {output_path}")
print(f"  Sheets   : SLA Model | Revenue Scenario Grid |"
      f" Churn Summary | Retention Dashboard")
print("\n✓ Step 7 complete — Ready for Power BI (Step 8)")